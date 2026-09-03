#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel 导入/导出通用工具 —— FastAPI Response + openpyxl"""

from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi.responses import StreamingResponse

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook, Workbook


def export_to_excel(data: List[Dict[str, Any]], sheet_name: str = "Sheet") -> bytes:
    """将字典列表导出为 .xlsx 二进制数据"""
    if not data:
        wb = Workbook()
        wb.active.append(["(空)"])
        return wb.export_as_bytes("xlsx") if hasattr(wb, 'export_as_bytes') else _save_xlsx(wb)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 表头
    headers = list(data[0].keys())
    ws.append(headers)

    # 数据行
    for row in data:
        ws.append([row.get(h, "") for h in headers])

    return _save_xlsx(wb)


def _save_xlsx(wb: Workbook) -> bytes:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def download_xlsx(data: List[Dict[str, Any]], filename: str, sheet_name: str = "Sheet") -> StreamingResponse:
    """返回一个 HTTP 下载响应"""
    raw = export_to_excel(data, sheet_name)
    return StreamingResponse(
        iter([raw]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def import_from_excel(file_content: bytes) -> Dict[str, Any]:
    """解析上传的 Excel，返回 {headers, rows}"""
    wb = load_workbook(filename=BytesIO(file_content))
    ws = wb.active
    if ws.max_row < 1:
        return {"headers": [], "rows": []}

    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                val = row[i]
                # Try to convert numeric strings
                if isinstance(val, str):
                    try:
                        val = int(val)
                    except ValueError:
                        try:
                            val = float(val)
                        except ValueError:
                            pass
                row_dict[header or f"col_{i}"] = val if val is not None else ""
            else:
                row_dict[header or f"col_{i}"] = ""
        rows.append(row_dict)

    return {"headers": headers, "rows": rows}
